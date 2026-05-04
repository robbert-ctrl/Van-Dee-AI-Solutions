#!/usr/bin/env python3
"""
Build a styled Excel workbook for the Tech Stack Audit lead magnet.

Output: public/tools/tech-stack-audit.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

# ─────── Brand colors ───────
NAVY = "0F172A"
ORANGE = "EA580C"
ORANGE_LIGHT = "FED7AA"
CREAM = "FAFAF7"
CREAM_DEEP = "F5F5EE"
INK_500 = "475569"
INK_400 = "94A3B8"
LINE = "E2E8F0"
EMERALD = "059669"
EMERALD_LIGHT = "D1FAE5"
RED = "DC2626"
RED_LIGHT = "FEE2E2"
YELLOW_LIGHT = "FEF3C7"

THIN = Side(border_style="thin", color=LINE)
NO_BORDER = Border()
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM = Border(bottom=THIN)
BORDER_BOLD_BOTTOM = Border(bottom=Side(border_style="medium", color=NAVY))


def style_header(cell, fill_color=NAVY, font_color="FFFFFF", size=11):
    cell.font = Font(name="Calibri", bold=True, color=font_color, size=size)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER_ALL


def style_title(cell, size=22, color=NAVY):
    cell.font = Font(name="Calibri", bold=True, size=size, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_subtitle(cell, size=11, color=INK_500):
    cell.font = Font(name="Calibri", size=size, color=color, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_label(cell, color=INK_500, bold=False, size=11):
    cell.font = Font(name="Calibri", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_data(cell, color="000000", size=11, align="left", bold=False):
    cell.font = Font(name="Calibri", size=size, color=color, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = Border(bottom=THIN)


def style_kpi(cell, color=NAVY, size=14, bold=True, align="right"):
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")


def style_section_band(cell, fill=ORANGE, font_color="FFFFFF"):
    cell.font = Font(name="Calibri", bold=True, size=12, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def fill_row(ws, row, start_col, end_col, color):
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)


def set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ─────── Build workbook ───────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ============================================================
# SHEET 1 — INSTRUCTIES
# ============================================================
ws = wb.create_sheet("Instructies")
ws.sheet_view.showGridLines = False
ws.sheet_view.tabColor = NAVY

# Column widths
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 2

# Title block
ws.merge_cells("B2:B2")
ws["B2"] = "Tech Stack Audit"
style_title(ws["B2"], size=28, color=NAVY)
set_row_height(ws, 2, 38)

ws["B3"] = "Van Dee AI Solutions  ·  Gratis template voor MKB-bedrijven"
style_subtitle(ws["B3"], size=12, color=ORANGE)
set_row_height(ws, 3, 22)

# Spacer
set_row_height(ws, 4, 10)

# Welkomstblok
ws["B5"] = "Welkom 👋"
style_label(ws["B5"], bold=True, size=14, color=NAVY)
set_row_height(ws, 5, 24)

ws["B6"] = ("Deze audit helpt je in 30 minuten te ontdekken welke tools je teveel betaalt, "
            "welke processen automatiseerbaar zijn met AI, en wat het je oplevert in euro's.")
ws["B6"].font = Font(name="Calibri", size=11, color=INK_500)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
set_row_height(ws, 6, 38)

set_row_height(ws, 7, 10)

# How-to
ws["B8"] = "Zo werkt het"
style_label(ws["B8"], bold=True, size=14, color=NAVY)
set_row_height(ws, 8, 24)

steps = [
    ("1", "Tabblad 'Tool Inventaris'", "Vul al je huidige software in. De aanbeveling-kolom geeft direct advies."),
    ("2", "Tabblad 'Gap Analyse'", "Welke processen kosten veel tijd? Welke zijn AI-rijp?"),
    ("3", "Tabblad 'ROI Calculator'", "Vul je uurtarief in. Zie meteen je jaarlijkse besparing."),
    ("4", "Tabblad 'Actieplan'", "Geprioriteerde lijst — waar begin je vandaag?"),
]

row = 9
for num, title, desc in steps:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
    ws.cell(row=row, column=2).value = f"  {num}    {title}"
    ws.cell(row=row, column=2).font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    ws.cell(row=row, column=2).alignment = Alignment(vertical="center")
    set_row_height(ws, row, 22)
    row += 1
    ws.cell(row=row, column=2).value = f"           {desc}"
    ws.cell(row=row, column=2).font = Font(name="Calibri", size=10, color=INK_500, italic=True)
    set_row_height(ws, row, 18)
    row += 1

set_row_height(ws, row, 14)
row += 1

# Tip box
ws.cell(row=row, column=2).value = "💡 Tip"
ws.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=12, color=ORANGE)
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=CREAM_DEEP)
set_row_height(ws, row, 22)
row += 1
ws.cell(row=row, column=2).value = ("De voorbeelddata in elk tabblad mag je gewoon vervangen door je eigen cijfers. "
                                    "De formules berekenen automatisch je totalen en besparingen.")
ws.cell(row=row, column=2).font = Font(name="Calibri", size=10, color=INK_500)
ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=CREAM_DEEP)
set_row_height(ws, row, 36)
row += 2

# CTA block
set_row_height(ws, row, 14)
row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
ws.cell(row=row, column=2).value = "Hulp nodig bij implementatie?"
ws.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws, row, 32)
row += 1
ws.cell(row=row, column=2).value = "Plan een gratis strategiegesprek (30 min, geen verkoopverhaal)"
ws.cell(row=row, column=2).font = Font(name="Calibri", size=11, color="FFFFFF")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws, row, 22)
row += 1
ws.cell(row=row, column=2).value = "👉  cal.com/vandeeaisolutions/discoverycall"
ws.cell(row=row, column=2).font = Font(name="Calibri", size=12, bold=True, color=ORANGE, underline="single")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=row, column=2).hyperlink = "https://cal.com/vandeeaisolutions/discoverycall"
set_row_height(ws, row, 28)


# ============================================================
# SHEET 2 — TOOL INVENTARIS
# ============================================================
ws = wb.create_sheet("1. Tool Inventaris")
ws.sheet_view.showGridLines = False
ws.sheet_view.tabColor = ORANGE

# Column widths
widths = [2, 5, 28, 22, 14, 12, 12, 12, 16, 28, 2]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title
ws["B2"] = "Tool Inventaris"
style_title(ws["B2"], size=22)
set_row_height(ws, 2, 32)
ws["B3"] = "Overzicht van alle tools die je nu gebruikt — vul je eigen cijfers in de groene kolommen in"
style_subtitle(ws["B3"])
set_row_height(ws, 3, 18)
set_row_height(ws, 4, 8)

# Headers
header_row = 5
headers = ["#", "Tool", "Categorie", "Kosten/maand (€)", "Gebruikers", "Gebruik (1-5)", "Waarde (1-5)", "€/gebruiker", "Aanbeveling"]
for i, h in enumerate(headers, 2):
    cell = ws.cell(row=header_row, column=i, value=h)
    style_header(cell, fill_color=NAVY)
set_row_height(ws, header_row, 28)

# Data rows
data = [
    (1, "Slack", "Communicatie", 120, 15, 5, 4, "HOUDEN"),
    (2, "HubSpot CRM", "Sales / Marketing", 450, 8, 4, 4, "HOUDEN"),
    (3, "Asana", "Project Management", 180, 12, 3, 3, "EVALUEREN"),
    (4, "Notion", "Documentatie", 80, 20, 4, 5, "HOUDEN"),
    (5, "Zapier", "Automatisering", 290, 3, 3, 3, "VERVANGEN"),
    (6, "Make.com", "Automatisering", 99, 3, 4, 5, "HOUDEN"),
    (7, "Claude Pro", "AI", 60, 3, 5, 5, "HOUDEN"),
    (8, "ChatGPT Plus", "AI", 60, 3, 2, 3, "SCHRAPPEN"),
    (9, "Google Workspace", "Productiviteit", 180, 15, 5, 5, "HOUDEN"),
    (10, "Microsoft 365", "Productiviteit", 225, 15, 3, 3, "EVALUEREN"),
]

start_data = header_row + 1
for idx, (num, tool, cat, cost, users, use, value, advice) in enumerate(data):
    r = start_data + idx
    cells = [num, tool, cat, cost, users, use, value, f"=D{r}/E{r}", advice]
    for i, val in enumerate(cells, 2):
        cell = ws.cell(row=r, column=i, value=val)
        cell.font = Font(name="Calibri", size=11, color="000000")
        cell.alignment = Alignment(horizontal="left" if i in (2, 3, 4) else "right", vertical="center")
        cell.border = Border(bottom=THIN)
        if i == 5:  # cost EUR
            cell.number_format = '€ #,##0'
        elif i == 9:  # cost per user
            cell.number_format = '€ #,##0.00'
        elif i == 10:  # advice
            cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # Alternating row color
    if idx % 2 == 1:
        fill_row(ws, r, 2, 10, CREAM_DEEP)
    set_row_height(ws, r, 22)

# Empty rows for user input
empty_start = start_data + len(data)
for i in range(5):
    r = empty_start + i
    num = len(data) + 1 + i
    ws.cell(row=r, column=2, value=num).font = Font(name="Calibri", size=11, color=INK_400, italic=True)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=r, column=2).border = Border(bottom=THIN)
    for col in range(3, 11):
        cell = ws.cell(row=r, column=col)
        cell.fill = PatternFill("solid", fgColor=EMERALD_LIGHT)
        cell.border = Border(bottom=THIN)
    set_row_height(ws, r, 22)

# Total row
total_row = empty_start + 5
data_end = empty_start + 4  # last empty row
for col in range(2, 11):
    ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=total_row, column=col).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(row=total_row, column=col).alignment = Alignment(vertical="center")
ws.cell(row=total_row, column=2, value="TOTAAL").alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.cell(row=total_row, column=2).font = Font(bold=True, color="FFFFFF", size=12)
ws.cell(row=total_row, column=5, value=f"=SUM(E{start_data}:E{data_end})").number_format = '€ #,##0'
ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right", vertical="center")
set_row_height(ws, total_row, 30)

# Annual summary
year_row = total_row + 2
ws.cell(row=year_row, column=2, value="Jaarlijkse kosten").font = Font(bold=True, size=12, color=INK_500)
ws.cell(row=year_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=year_row, column=5, value=f"=E{total_row}*12").number_format = '€ #,##0'
ws.cell(row=year_row, column=5).font = Font(bold=True, size=14, color=NAVY)
ws.cell(row=year_row, column=5).alignment = Alignment(horizontal="right", vertical="center")
set_row_height(ws, year_row, 26)

# Conditional formatting on Aanbeveling column (J)
cf_range = f"J{start_data}:J{data_end}"
green_fill = PatternFill("solid", fgColor=EMERALD_LIGHT)
yellow_fill = PatternFill("solid", fgColor=YELLOW_LIGHT)
red_fill = PatternFill("solid", fgColor=RED_LIGHT)
ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f'J{start_data}="HOUDEN"'], fill=green_fill, font=Font(bold=True, color=EMERALD)))
ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f'J{start_data}="EVALUEREN"'], fill=yellow_fill, font=Font(bold=True, color="92400E")))
ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f'OR(J{start_data}="VERVANGEN",J{start_data}="SCHRAPPEN")'], fill=red_fill, font=Font(bold=True, color=RED)))

# Color scale on Use & Value columns (G and H)
color_scale = ColorScaleRule(
    start_type="num", start_value=1, start_color="FECACA",
    mid_type="num", mid_value=3, mid_color="FEF3C7",
    end_type="num", end_value=5, end_color="A7F3D0"
)
ws.conditional_formatting.add(f"G{start_data}:H{data_end}", color_scale)

# Legend
legend_row = year_row + 3
ws.cell(row=legend_row, column=2, value="Legenda scoring").font = Font(bold=True, size=11, color=INK_500)
set_row_height(ws, legend_row, 22)
legend_row += 1
ws.cell(row=legend_row, column=2, value="Gebruik (1-5)").font = Font(size=10, color=INK_500, italic=True)
ws.cell(row=legend_row, column=3, value="1 = Nauwelijks  ·  3 = Wekelijks  ·  5 = Continu dagelijks").font = Font(size=10, color=INK_500)
legend_row += 1
ws.cell(row=legend_row, column=2, value="Waarde (1-5)").font = Font(size=10, color=INK_500, italic=True)
ws.cell(row=legend_row, column=3, value="1 = Onmisbaar  ·  3 = Handig  ·  5 = Kan zonder").font = Font(size=10, color=INK_500)


# ============================================================
# SHEET 3 — GAP ANALYSE
# ============================================================
ws = wb.create_sheet("2. Gap Analyse")
ws.sheet_view.showGridLines = False
ws.sheet_view.tabColor = ORANGE

widths = [2, 5, 32, 22, 14, 18, 18, 32, 2]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws["B2"] = "Gap Analyse"
style_title(ws["B2"], size=22)
set_row_height(ws, 2, 32)
ws["B3"] = "Welke processen kosten te veel tijd? Welke zijn geschikt voor AI-automatisering?"
style_subtitle(ws["B3"])
set_row_height(ws, 3, 18)
set_row_height(ws, 4, 8)

headers = ["#", "Proces", "Huidige tool", "Uren/week", "Automatiseerbaar", "Besparing (u/wk)", "AI-oplossing"]
for i, h in enumerate(headers, 2):
    cell = ws.cell(row=5, column=i, value=h)
    style_header(cell, fill_color=NAVY)
set_row_height(ws, 5, 28)

processes = [
    ("Facturen verwerken", "Excel + Exact", 10, "JA", 8, "GPT-4 Vision + Make.com + Exact API"),
    ("Email sorteren / beantwoorden", "Outlook handmatig", 8, "DEELS", 5, "Claude + email-filters"),
    ("Klantvragen eerste lijn", "Telefoon + email", 12, "DEELS", 6, "AI Chatbot op website"),
    ("Wekelijkse rapportages", "Excel handmatig", 5, "JA", 4, "Claude + Google Sheets API"),
    ("LinkedIn content schrijven", "CEO eigen tijd", 3, "JA", 2, "Claude + content-templates"),
    ("Data-entry uit formulieren", "Handmatig typen", 6, "JA", 5, "GPT-4 Vision + Airtable"),
]

start_data = 6
for idx, (proc, tool, hours, auto, save, ai) in enumerate(processes):
    r = start_data + idx
    ws.cell(row=r, column=2, value=idx + 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=r, column=3, value=proc)
    ws.cell(row=r, column=4, value=tool)
    ws.cell(row=r, column=5, value=hours).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=r, column=5).number_format = "0"
    ws.cell(row=r, column=6, value=auto).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=6).font = Font(bold=True, color=NAVY, size=11)
    ws.cell(row=r, column=7, value=save).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=r, column=7).number_format = "0"
    ws.cell(row=r, column=8, value=ai).font = Font(size=10, color=INK_500, italic=True)
    for col in range(2, 9):
        ws.cell(row=r, column=col).border = Border(bottom=THIN)
        if not ws.cell(row=r, column=col).font.size:
            ws.cell(row=r, column=col).font = Font(name="Calibri", size=11)
        ws.cell(row=r, column=col).alignment = Alignment(
            horizontal=ws.cell(row=r, column=col).alignment.horizontal or "left",
            vertical="center"
        )
    if idx % 2 == 1:
        fill_row(ws, r, 2, 8, CREAM_DEEP)
    set_row_height(ws, r, 24)

# Conditional formatting on Automatiseerbaar
cf_range = f"F{start_data}:F{start_data + len(processes) - 1}"
ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f'F{start_data}="JA"'], fill=PatternFill("solid", fgColor=EMERALD_LIGHT), font=Font(bold=True, color=EMERALD)))
ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f'F{start_data}="DEELS"'], fill=PatternFill("solid", fgColor=YELLOW_LIGHT), font=Font(bold=True, color="92400E")))
ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f'F{start_data}="NEE"'], fill=PatternFill("solid", fgColor=RED_LIGHT), font=Font(bold=True, color=RED)))

# Empty rows
empty_start = start_data + len(processes)
for i in range(4):
    r = empty_start + i
    ws.cell(row=r, column=2, value=len(processes) + 1 + i).font = Font(name="Calibri", size=11, color=INK_400, italic=True)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=r, column=2).border = Border(bottom=THIN)
    for col in range(3, 9):
        cell = ws.cell(row=r, column=col)
        cell.fill = PatternFill("solid", fgColor=EMERALD_LIGHT)
        cell.border = Border(bottom=THIN)
    set_row_height(ws, r, 22)

# Totals
total_row = empty_start + 4
data_end = empty_start + 3
for col in range(2, 9):
    ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=total_row, column=col).font = Font(bold=True, color="FFFFFF", size=12)
ws.cell(row=total_row, column=2, value="TOTAAL").alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.cell(row=total_row, column=5, value=f"=SUM(E{start_data}:E{data_end})").alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=total_row, column=5).number_format = "0\" uren\""
ws.cell(row=total_row, column=7, value=f"=SUM(G{start_data}:G{data_end})").alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=total_row, column=7).number_format = "0\" uren\""
set_row_height(ws, total_row, 30)

# Highlight summary
summary_row = total_row + 2
ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=8)
ws.cell(row=summary_row, column=2, value=f"Totale potentiële tijdsbesparing per week — gebruik dit getal in de ROI Calculator")
ws.cell(row=summary_row, column=2).font = Font(bold=True, size=11, color=ORANGE, italic=True)
ws.cell(row=summary_row, column=2).fill = PatternFill("solid", fgColor=ORANGE_LIGHT)
ws.cell(row=summary_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws, summary_row, 28)


# ============================================================
# SHEET 4 — ROI CALCULATOR
# ============================================================
ws = wb.create_sheet("3. ROI Calculator")
ws.sheet_view.showGridLines = False
ws.sheet_view.tabColor = EMERALD

widths = [2, 36, 22, 22, 2]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws["B2"] = "ROI Calculator"
style_title(ws["B2"], size=22)
set_row_height(ws, 2, 32)
ws["B3"] = "Vul je uurtarief en geschatte besparing in — zie meteen wat het oplevert"
style_subtitle(ws["B3"])
set_row_height(ws, 3, 18)
set_row_height(ws, 4, 12)

# Inputs section
ws.merge_cells("B5:D5")
ws["B5"] = "  INVOER  (vul deze in)"
style_section_band(ws["B5"], fill=NAVY, font_color="FFFFFF")
set_row_height(ws, 5, 28)

inputs = [
    ("Gemiddeld uurtarief in jouw bedrijf", 50, '€ #,##0'),
    ("Geschatte besparing per week (uren)", 20, '0" uren"'),
    ("Aantal werkweken per jaar", 50, '0" weken"'),
]

input_rows = []
for idx, (label, value, fmt) in enumerate(inputs):
    r = 6 + idx
    ws.cell(row=r, column=2, value=label).font = Font(name="Calibri", size=11, color=INK_500)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=r, column=2).border = Border(bottom=THIN)
    cell = ws.cell(row=r, column=4, value=value)
    cell.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = PatternFill("solid", fgColor=EMERALD_LIGHT)
    cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    cell.number_format = fmt
    set_row_height(ws, r, 28)
    input_rows.append(r)

set_row_height(ws, 9, 12)

# Outputs section
ws.merge_cells("B10:D10")
ws["B10"] = "  RESULTATEN  (automatisch berekend)"
style_section_band(ws["B10"], fill=ORANGE, font_color="FFFFFF")
set_row_height(ws, 10, 28)

# Output rows with formulas
output_data = [
    ("Besparing per week", f"=D{input_rows[0]}*D{input_rows[1]}", '€ #,##0', NAVY, 12),
    ("Besparing per maand", f"=D{input_rows[0]}*D{input_rows[1]}*4.33", '€ #,##0', NAVY, 12),
    ("Besparing per jaar", f"=D{input_rows[0]}*D{input_rows[1]}*D{input_rows[2]}", '€ #,##0', ORANGE, 18),
]

for idx, (label, formula, fmt, color, size) in enumerate(output_data):
    r = 11 + idx
    ws.cell(row=r, column=2, value=label).font = Font(name="Calibri", size=11, color=INK_500)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center", indent=1)
    ws.cell(row=r, column=2).border = Border(bottom=THIN)
    cell = ws.cell(row=r, column=4, value=formula)
    cell.font = Font(name="Calibri", size=size, bold=True, color=color)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = Border(bottom=THIN)
    cell.number_format = fmt
    set_row_height(ws, r, 32 if size > 14 else 28)

set_row_height(ws, 14, 16)

# Investment section
ws.merge_cells("B15:D15")
ws["B15"] = "  PROJECT INVESTERING & TERUGVERDIENTIJD"
style_section_band(ws["B15"], fill=INK_500, font_color="FFFFFF")
set_row_height(ws, 15, 28)

inv_rows = [
    ("Eenmalige AI-implementatie investering (gemiddeld)", 12000, '€ #,##0'),
]
inv_input_row = 16
ws.cell(row=inv_input_row, column=2, value=inv_rows[0][0]).font = Font(name="Calibri", size=11, color=INK_500)
ws.cell(row=inv_input_row, column=2).alignment = Alignment(vertical="center", indent=1)
ws.cell(row=inv_input_row, column=2).border = Border(bottom=THIN)
cell = ws.cell(row=inv_input_row, column=4, value=inv_rows[0][1])
cell.font = Font(size=12, bold=True, color=NAVY)
cell.alignment = Alignment(horizontal="right", vertical="center")
cell.fill = PatternFill("solid", fgColor=EMERALD_LIGHT)
cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
cell.number_format = '€ #,##0'
set_row_height(ws, inv_input_row, 28)

ws.cell(row=17, column=2, value="Terugverdientijd").font = Font(name="Calibri", size=11, color=INK_500)
ws.cell(row=17, column=2).alignment = Alignment(vertical="center", indent=1)
ws.cell(row=17, column=2).border = Border(bottom=THIN)
ws.cell(row=17, column=4, value=f"=D{inv_input_row}/(D{input_rows[0]}*D{input_rows[1]}*4.33)")
ws.cell(row=17, column=4).font = Font(size=14, bold=True, color=NAVY)
ws.cell(row=17, column=4).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=17, column=4).number_format = '0.0" maanden"'
ws.cell(row=17, column=4).border = Border(bottom=THIN)
set_row_height(ws, 17, 28)

ws.cell(row=18, column=2, value="ROI in jaar 1").font = Font(name="Calibri", size=11, color=INK_500)
ws.cell(row=18, column=2).alignment = Alignment(vertical="center", indent=1)
ws.cell(row=18, column=2).border = Border(bottom=THIN)
ws.cell(row=18, column=4, value=f"=(D{input_rows[0]}*D{input_rows[1]}*D{input_rows[2]}-D{inv_input_row})/D{inv_input_row}")
ws.cell(row=18, column=4).font = Font(size=18, bold=True, color=ORANGE)
ws.cell(row=18, column=4).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=18, column=4).number_format = '0%'
ws.cell(row=18, column=4).border = Border(bottom=THIN)
set_row_height(ws, 18, 32)

# Note
set_row_height(ws, 20, 12)
ws.merge_cells("B21:D21")
ws["B21"] = ("💡 De getallen hierboven zijn voorbeeld-input. Vervang ze door je eigen cijfers — de formules berekenen automatisch je ROI.")
ws["B21"].font = Font(name="Calibri", size=10, color=INK_500, italic=True)
ws["B21"].fill = PatternFill("solid", fgColor=CREAM_DEEP)
ws["B21"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)
set_row_height(ws, 21, 36)


# ============================================================
# SHEET 5 — ACTIEPLAN
# ============================================================
ws = wb.create_sheet("4. Actieplan")
ws.sheet_view.showGridLines = False
ws.sheet_view.tabColor = ORANGE

widths = [2, 8, 36, 16, 18, 14, 14, 2]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws["B2"] = "Prioriteiten Actieplan"
style_title(ws["B2"], size=22)
set_row_height(ws, 2, 32)
ws["B3"] = "Geprioriteerd op impact (uren bespaard) ÷ implementatietijd — hoogste score eerst"
style_subtitle(ws["B3"])
set_row_height(ws, 3, 18)
set_row_height(ws, 4, 8)

headers = ["Prioriteit", "Actie", "Impact (u/wk)", "Implementatietijd", "Score", "Start"]
for i, h in enumerate(headers, 2):
    cell = ws.cell(row=5, column=i, value=h)
    style_header(cell)
set_row_height(ws, 5, 28)

actions = [
    (1, "Email classificatie & routing", 5, "2 weken", "Maand 1"),
    (2, "Factuurverwerking automatiseren", 8, "4 weken", "Maand 1"),
    (3, "AI chatbot eerste lijn klantvragen", 6, "3 weken", "Maand 1"),
    (4, "Wekelijkse rapportages automatiseren", 4, "2 weken", "Maand 2"),
    (5, "Content-generator LinkedIn", 2, "1 week", "Maand 2"),
    (6, "Data-entry uit formulieren", 5, "3 weken", "Maand 3"),
]

start_data = 6
for idx, (prio, action, impact, time, start) in enumerate(actions):
    r = start_data + idx
    # Priority badge
    cell = ws.cell(row=r, column=2, value=prio)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor=ORANGE if prio <= 3 else INK_400)
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    # Action
    ws.cell(row=r, column=3, value=action).font = Font(name="Calibri", size=11, bold=(prio <= 3), color=NAVY)
    ws.cell(row=r, column=3).alignment = Alignment(vertical="center", indent=1)
    # Impact
    ws.cell(row=r, column=4, value=impact).number_format = '0" uren"'
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=r, column=4).font = Font(size=11, color=INK_500)
    # Time
    ws.cell(row=r, column=5, value=time).font = Font(name="Calibri", size=11, color=INK_500)
    ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
    # Score (computed)
    weeks = int(time.split()[0])
    score = round(impact / weeks, 2)
    ws.cell(row=r, column=6, value=score).font = Font(bold=True, size=12, color=NAVY if prio > 3 else ORANGE)
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=6).number_format = '0.00'
    # Start
    ws.cell(row=r, column=7, value=start).font = Font(name="Calibri", size=11, color=INK_500, italic=True)
    ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")
    for col in range(3, 8):
        ws.cell(row=r, column=col).border = Border(bottom=THIN)
    set_row_height(ws, r, 32)

# Tip box at bottom
tip_row = start_data + len(actions) + 2
ws.merge_cells(start_row=tip_row, start_column=2, end_row=tip_row, end_column=7)
ws.cell(row=tip_row, column=2, value="💡 Begin met Prioriteit 1, 2 en 3 — die hebben de hoogste score (impact ÷ tijd) en daarmee de snelste payback")
ws.cell(row=tip_row, column=2).font = Font(name="Calibri", size=11, bold=True, color=ORANGE, italic=True)
ws.cell(row=tip_row, column=2).fill = PatternFill("solid", fgColor=ORANGE_LIGHT)
ws.cell(row=tip_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws, tip_row, 36)


# ============================================================
# SHEET 6 — VOLGENDE STAPPEN
# ============================================================
ws = wb.create_sheet("5. Volgende Stappen")
ws.sheet_view.showGridLines = False
ws.sheet_view.tabColor = NAVY

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 90

ws["B2"] = "Volgende Stappen"
style_title(ws["B2"], size=22)
set_row_height(ws, 2, 32)

ws["B3"] = "Je hebt nu een complete audit. Wat doe je hiermee?"
style_subtitle(ws["B3"], size=12)
set_row_height(ws, 3, 22)
set_row_height(ws, 4, 12)

# What you got
ws["B5"] = "Wat je nu hebt"
style_label(ws["B5"], bold=True, size=14, color=NAVY)
set_row_height(ws, 5, 24)

what_got = [
    "✓  Compleet overzicht van alle huidige tools en jaarlijkse kosten",
    "✓  Inzicht in welke tools je kunt schrappen of vervangen",
    "✓  Concrete AI-oplossingen voor je bedrijfsprocessen",
    "✓  ROI-berekening op basis van je eigen cijfers",
    "✓  Geprioriteerde implementatie-volgorde",
]
row = 6
for item in what_got:
    ws.cell(row=row, column=2, value=item).font = Font(name="Calibri", size=11, color=NAVY)
    ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=2)
    set_row_height(ws, row, 24)
    row += 1

set_row_height(ws, row, 16)
row += 1

# Two paths
ws["B" + str(row)] = "Twee paden vanaf hier"
style_label(ws["B" + str(row)], bold=True, size=14, color=NAVY)
set_row_height(ws, row, 24)
row += 1

paths = [
    ("Pad 1 — Doe het zelf", "Begin met Prioriteit 1 uit je actieplan. Gebruik tools als Make.com, OpenAI API en Google Sheets. Reken op 2-4 weken implementatietijd per proces.", INK_500),
    ("Pad 2 — Laat het bouwen", "Boek een gratis strategiegesprek (30 min). Wij analyseren je audit, geven onafhankelijk advies, en bouwen het indien gewenst voor je. Geen verkoopverhaal.", ORANGE),
]

for title, desc, color in paths:
    set_row_height(ws, row, 12)
    row += 1
    ws.cell(row=row, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=color)
    ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)
    set_row_height(ws, row, 24)
    row += 1
    ws.cell(row=row, column=2, value=desc).font = Font(name="Calibri", size=11, color=INK_500)
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    set_row_height(ws, row, 50)
    row += 1

# CTA block
set_row_height(ws, row, 16)
row += 1

ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
ws.cell(row=row, column=2, value="Plan je gratis strategiegesprek").font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws, row, 44)
row += 1

ws.cell(row=row, column=2, value="30 minuten · Vrijblijvend · Geen verkoopverhaal").font = Font(name="Calibri", size=11, color="FFFFFF")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws, row, 24)
row += 1

ws.cell(row=row, column=2, value="👉  cal.com/vandeeaisolutions/discoverycall").font = Font(name="Calibri", size=14, bold=True, color=ORANGE, underline="single")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=row, column=2).hyperlink = "https://cal.com/vandeeaisolutions/discoverycall"
set_row_height(ws, row, 32)
row += 1

set_row_height(ws, row, 12)
row += 1

# Contact info
ws.cell(row=row, column=2, value="Robbert van Dee  ·  Founder Van Dee AI Solutions").font = Font(name="Calibri", bold=True, size=11, color=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws, row, 22)
row += 1

ws.cell(row=row, column=2, value="Tiel, Gelderland  ·  KvK 85394041  ·  robbert@vandeeaisolutions.com").font = Font(name="Calibri", size=10, color=INK_500)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws, row, 20)
row += 1

ws.cell(row=row, column=2, value="vandeeaisolutions.com").font = Font(name="Calibri", size=10, color=ORANGE, underline="single")
ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=row, column=2).hyperlink = "https://vandeeaisolutions.com"
set_row_height(ws, row, 20)


# ─────── Save ───────
output_path = Path(__file__).parent.parent / "public" / "tools" / "tech-stack-audit.xlsx"
output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(output_path)
print(f"✅ Saved: {output_path}")
print(f"   Size: {output_path.stat().st_size / 1024:.1f} KB")
print(f"   Sheets: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
